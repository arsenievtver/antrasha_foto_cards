# -*- coding: utf-8 -*-
"""
Заливка заказов SS27 в procurement API.

Использование:
  # dry-run (только сверка справочников, без POST)
  ANTRASHA_API=https://new.antrasha.ru \\
  ANTRASHA_PHONE='+7...' ANTRASHA_PIN='......' \\
  python3 .tmp_ms/upload_orders_ss27.py

  # реальная заливка
  ... python3 .tmp_ms/upload_orders_ss27.py --apply

Логин: POST /auth/login (worker с правом product) или
  ANTRASHA_TOKEN='...' без phone/pin.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DETAIL_JSON = ROOT / ".tmp_ms" / "SS27_заказы_детали.json"
# Источник истины — Excel «Сводка» (туда вручную правят суммы).
DEFAULT_XLSX = Path("/Users/alekseiarsenev/Downloads/заказы/SS27_заказы_по_категориям.xlsx")
FALLBACK_XLSX = ROOT / ".tmp_ms" / "SS27_заказы_по_категориям.xlsx"

# moy_sklad_id / имена из procurementCategories.js → ищем в refs
CAT_MS_IDS = {
    "Верхняя одежда муж": ["0ebca617-f97a-11e9-0a80-0579004f6022"],
    "Пиджаки, жакеты, бомбер муж": ["009bd151-b37b-11e9-9ff4-3150003a1bb1"],
    "Футболки, поло муж": ["46a5c5b7-5708-11e9-9ff4-315000d0798d"],
    "Брюки, джинсы муж": ["46b4f0d3-5708-11e9-9ff4-315000d079ad"],
    "Бриджи, шорты муж": ["55edd126-8bff-11f1-0a80-142f000aee50"],
    "Трикотаж муж": ["7958c78e-9e44-11e9-9ff4-31500007d713"],
    "Рубашки": ["797d0e35-9e44-11e9-9ff4-31500007d733"],
    "Костюмы муж": ["eec41100-9847-11eb-0a80-0616000ac009"],
    "Обувь муж": ["f8fae156-b37a-11e9-9ff4-3150003a11ec"],
    "Аксессуары муж": ["82adf299-8e8b-11e9-9ff4-31500007fc47"],
    "Верхняя одежда жен": ["0dea4445-f97a-11e9-0a80-0579004f5ecf"],
    "Пиджаки, жакеты, бомбер жен": [
        "79292943-9e44-11e9-9ff4-31500007d6f3",
        "463e7bec-34dd-11f1-0a80-148d00118078",
    ],
    "Футболки, поло, топы жен": ["f7b6946e-b37a-11e9-9ff4-3150003a0ff5"],
    "Блузки, рубашки жен": ["21e1d207-b53f-11e9-9ff4-31500015315b"],
    "Трикотаж жен": ["cd27a401-d3a6-11e9-0a80-02690003e199"],
    "Брюки, джинсы жен": [
        "78fabba1-9e44-11e9-9ff4-31500007d6c1",
        "8ade28c6-6e3e-11f1-0a80-00b0001171b1",
    ],
    "Бриджи, шорты жен": ["4643b20e-8bfa-11f1-0a80-18830009f9ac"],
    "Платья жен": ["65dca14b-8bfd-11f1-0a80-0fbf000a6721"],
    "Юбки жен": ["26114fa1-a495-11e9-9ff4-3150000fa9a1"],
    "Обувь жен": ["79419e87-9e44-11e9-9ff4-31500007d6fe"],
    "Аксессуары жен": [],
}

BRAND_ALIASES = {
    "Aeronautica": ["aeronautica", "aeronautica militare"],
    "Diktat": ["diktat"],
    "Duno": ["duno", "d-uno", "d uno"],
    "Gardeur": ["gardeur"],
    "RIANI": ["riani"],
    "Roy Robson": ["roy robson", "royrobson"],
    "SeidenSticker": ["seidensticker", "seiden sticker"],
    "TRANSIT": ["transit"],
}

GENDER_MAP = {"муж": "men", "жен": "women"}


def api_base() -> str:
    base = (os.environ.get("ANTRASHA_API") or "https://new.antrasha.ru").rstrip("/")
    return base


def http_json(method: str, path: str, token: str | None = None, body: dict | None = None):
    base = api_base()
    # nginx may serve API at /api/*
    urls = []
    if path.startswith("/api/"):
        urls.append(base + path)
    else:
        urls.append(base + "/api" + path)
        urls.append(base + path)

    data = None if body is None else json.dumps(body).encode("utf-8")
    last_err = None
    for url in urls:
        headers = {"Content-Type": "application/json", "Accept": "application/json"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                raw = resp.read().decode("utf-8")
                if not raw.strip():
                    return {}
                return json.loads(raw)
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            last_err = f"HTTP {e.code} {url}: {err_body[:500]}"
            # 404 on /api prefix → try next
            if e.code == 404:
                continue
            raise SystemExit(last_err) from e
        except Exception as e:
            last_err = f"{url}: {e}"
            continue
    raise SystemExit(last_err or "request failed")


def login() -> str:
    token = os.environ.get("ANTRASHA_TOKEN", "").strip()
    if token:
        return token
    phone = os.environ.get("ANTRASHA_PHONE", "").strip()
    pin = os.environ.get("ANTRASHA_PIN", "").strip()
    if not phone or not pin:
        raise SystemExit(
            "Задайте ANTRASHA_PHONE + ANTRASHA_PIN (worker с правом product)\n"
            "или ANTRASHA_TOKEN=... (Bearer)"
        )
    # normalize RU phone to +7XXXXXXXXXX
    digits = "".join(ch for ch in phone if ch.isdigit())
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10 and digits.startswith("9"):
        digits = "7" + digits
    if len(digits) == 11 and digits.startswith("7"):
        phone = "+" + digits
    data = http_json("POST", "/auth/login", body={"phone": phone, "pin": pin})
    role = data.get("role")
    perms = data.get("permissions") or []
    if role not in ("worker", "superuser"):
        raise SystemExit(f"Нет доступа: role={role}")
    if role == "worker" and "product" not in perms and role != "superuser":
        raise SystemExit(f"Нет права product. permissions={perms}")
    print(f"OK login role={role} perms={perms}")
    return data["access_token"]


def find_season(seasons: list[dict]) -> dict:
    keys = ("весна-лето 2027", "весна лето 2027", "vl2027", "ss27", "ss 27", "вл2027", "вл 2027")
    for s in seasons:
        blob = f"{s.get('name','')} {s.get('code','')}".lower().replace("ё", "е")
        if any(k in blob for k in keys):
            return s
    # fallback: print all
    print("Сезоны в системе:")
    for s in seasons:
        print(f"  {s.get('id')} | {s.get('code')} | {s.get('name')}")
    raise SystemExit("Не найден сезон Весна-лето 2027 — поправьте SEASON_ID вручную")


def find_brand(brands: list[dict], our_name: str) -> dict | None:
    aliases = BRAND_ALIASES.get(our_name, [our_name.lower()])
    # Prefer exact (case-insensitive), then alias containment — never match junk "-"
    candidates = [b for b in brands if (b.get("name") or "").strip() not in ("", "-")]
    our_l = our_name.lower()
    for b in candidates:
        if (b.get("name") or "").lower() == our_l:
            return b
    for b in candidates:
        bn = (b.get("name") or "").lower()
        for a in aliases:
            if len(a) < 3:
                continue
            if a == bn or a in bn or bn == a:
                return b
    return None


def resolve_category_id(categories: list[dict], label: str) -> str | None:
    by_ms = {}
    by_name = {}
    for c in categories:
        if c.get("moy_sklad_id"):
            by_ms[c["moy_sklad_id"]] = c
        by_name[c.get("name") or ""] = c

    for ms in CAT_MS_IDS.get(label, []):
        if ms in by_ms:
            return by_ms[ms]["id"]
    if label in by_name:
        return by_name[label]["id"]
    # alias names from CATEGORY_RULES
    aliases = {
        "Брюки, джинсы жен": [
            "Брюки, джинсы, бриджи, шорты жен",
            "Брюки, джинсы, бриджи, шорты",
        ],
        "Пиджаки, жакеты, бомбер жен": ["Пиджаки, жакеты, бомбер"],
        "Блузки, рубашки жен": ["Блузки, рубашки"],
        "Платья жен": ["Платья"],
        "Юбки жен": ["Юбки", "Платья, юбки", "Платья, юбки жен"],
        "Аксессуары муж": ["Аксессуары"],
        "Аксессуары жен": ["Аксессуары"],
    }
    for alt in aliases.get(label, []):
        if alt in by_name:
            return by_name[alt]["id"]
    for name, c in by_name.items():
        if label.lower() in name.lower() or name.lower() in label.lower():
            return c["id"]
    return None


def ensure_women_core_categories(token: str) -> None:
    """Женские брюки/пиджаки на проде иногда is_active=false — включаем для форм и импорта."""
    data = http_json("GET", "/admin/categories", token=token)
    items = data.get("items") or []
    need = {
        "Брюки, джинсы жен",
        "Пиджаки, жакеты, бомбер жен",
    }
    for c in items:
        name = c.get("name") or ""
        if name in need and c.get("is_active") is False:
            print(f"Activate category: {name} ({c['id']})")
            http_json(
                "PATCH",
                f"/admin/categories/{c['id']}",
                token=token,
                body={"is_active": True},
            )

def build_orders_from_json(detail: dict) -> dict[tuple[str, str], dict[str, float]]:
    matrix: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for ln in detail["lines"]:
        matrix[(ln["brand"], ln["gender"])][ln["category"]] += float(ln["amount"])
    return matrix


def parse_brand_gender(cell: str) -> tuple[str, str] | None:
    """'Duno муж' / 'Roy Robson муж' / 'ИТОГО' → (brand, муж|жен)."""
    s = (cell or "").strip()
    if not s or s.upper().startswith("ИТОГ"):
        return None
    if s.endswith(" муж"):
        return s[: -len(" муж")].strip(), "муж"
    if s.endswith(" жен"):
        return s[: -len(" жен")].strip(), "жен"
    return None


def build_orders_from_xlsx(path: Path) -> dict[tuple[str, str], dict[str, float]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "Сводка" not in wb.sheetnames:
        raise SystemExit(f"В {path} нет листа «Сводка»")
    ws = wb["Сводка"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Пустая сводка")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    matrix: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        parsed = parse_brand_gender(str(row[0]))
        if not parsed:
            continue
        brand, gender = parsed
        for i in range(1, len(header)):
            col = header[i]
            if not col or col == "Итого":
                continue
            val = row[i] if i < len(row) else None
            if val in (None, "", 0, 0.0):
                continue
            amount = round(float(val), 2)
            if amount > 0:
                matrix[(brand, gender)][col] += amount
    return matrix


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="Реально создать заказы")
    ap.add_argument("--ordered-on", default=str(date.today()))
    ap.add_argument("--comment-prefix", default="Импорт заказов SS27 из PDF/XLSX")
    ap.add_argument(
        "--xlsx",
        default=str(DEFAULT_XLSX if DEFAULT_XLSX.is_file() else FALLBACK_XLSX),
        help="Excel со сводкой (по умолчанию Downloads/заказы/...)",
    )
    ap.add_argument(
        "--from-json",
        action="store_true",
        help="Брать суммы из SS27_заказы_детали.json вместо Excel",
    )
    args = ap.parse_args()

    if args.from_json:
        detail = json.loads(DETAIL_JSON.read_text(encoding="utf-8"))
        matrix = build_orders_from_json(detail)
        print(f"Источник: JSON {DETAIL_JSON}")
    else:
        xlsx = Path(args.xlsx)
        if not xlsx.is_file():
            raise SystemExit(f"Нет файла: {xlsx}")
        matrix = build_orders_from_xlsx(xlsx)
        print(f"Источник: Excel {xlsx}")

    token = login()
    ensure_women_core_categories(token)
    refs = http_json("GET", "/admin/procurement/refs", token=token)
    # Refs only active — also merge full /admin/categories for resolution
    all_cats = http_json("GET", "/admin/categories", token=token).get("items") or []
    seasons = refs.get("seasons") or []
    brands = refs.get("brands") or []
    categories = all_cats or (refs.get("categories") or [])

    season = find_season(seasons)
    print(f"Сезон: {season.get('name')} ({season.get('code')}) id={season['id']}")

    # existing orders for season
    existing = http_json(
        "GET",
        f"/admin/brand-orders?season_id={season['id']}&limit=200",
        token=token,
    )
    existing_items = existing.get("items") or existing.get("orders") or []
    print(f"Уже заказов в сезоне: {len(existing_items)}")

    payloads = []
    missing_brands = []
    missing_cats = []
    for (brand, gender_ru), cats in sorted(matrix.items()):
        b = find_brand(brands, brand)
        if not b:
            missing_brands.append(brand)
            continue
        gender = GENDER_MAP[gender_ru]
        lines = []
        for label, amount in sorted(cats.items(), key=lambda x: -x[1]):
            amount = round(amount, 2)
            if amount <= 0:
                continue
            cid = resolve_category_id(categories, label)
            if not cid:
                missing_cats.append(label)
                continue
            lines.append(
                {
                    "category_id": cid,
                    "amount_eur": f"{amount:.2f}",
                    "comment": None,
                }
            )
        payloads.append(
            {
                "season_id": season["id"],
                "brand_id": b["id"],
                "brand_name": b["name"],
                "our_brand": brand,
                "gender": gender,
                "ordered_on": args.ordered_on,
                "has_prepayment": False,
                "comment": f"{args.comment_prefix}: {brand} {gender_ru}",
                "lines": lines,
                "total": round(sum(float(x["amount_eur"]) for x in lines), 2),
            }
        )

    if missing_brands:
        print("НЕ НАЙДЕНЫ БРЕНДЫ:", missing_brands)
        print("Бренды в системе (фрагмент):", [b["name"] for b in brands][:40])
    if missing_cats:
        print("НЕ НАЙДЕНЫ КАТЕГОРИИ:", sorted(set(missing_cats)))
        print("Категории:", [c["name"] for c in categories])

    print("\nПлан заливки:")
    for p in payloads:
        print(
            f"  {p['our_brand']} → {p['brand_name']} / {p['gender']}: "
            f"{p['total']:.2f} EUR, lines={len(p['lines'])}"
        )
    print(f"ИТОГО заказов: {len(payloads)}, сумма: {sum(p['total'] for p in payloads):.2f}")

    if missing_brands or missing_cats:
        raise SystemExit("Сначала заведите недостающие бренды/категории")

    if not args.apply:
        print("\nDry-run. Для заливки добавьте --apply")
        return

    created = []
    for p in payloads:
        body = {
            "season_id": p["season_id"],
            "brand_id": p["brand_id"],
            "gender": p["gender"],
            "ordered_on": p["ordered_on"],
            "has_prepayment": False,
            "comment": p["comment"],
            "lines": p["lines"],
        }
        row = http_json("POST", "/admin/brand-orders", token=token, body=body)
        created.append(row)
        print(
            f"CREATED {p['brand_name']} {p['gender']}: "
            f"id={row.get('id')} amount={row.get('amount_eur')}"
        )

    out = ROOT / ".tmp_ms" / "SS27_upload_result.json"
    out.write_text(json.dumps(created, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"\nГотово: {len(created)} заказов. Результат: {out}")


if __name__ == "__main__":
    main()
