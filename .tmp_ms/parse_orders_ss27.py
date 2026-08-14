# -*- coding: utf-8 -*-
"""Parse SS27 order PDFs/XLSX → Excel by brand × procurement categories."""
from __future__ import annotations

import json
import re
import subprocess
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ORDERS_DIR = Path("/Users/alekseiarsenev/Downloads/заказы")
OUT_XLSX = Path("/Users/alekseiarsenev/WebstormProjects/antrasha_tinder/.tmp_ms/SS27_заказы_по_категориям.xlsx")
OUT_DETAIL = Path("/Users/alekseiarsenev/WebstormProjects/antrasha_tinder/.tmp_ms/SS27_заказы_детали.json")
OUT_XLSX_DL = ORDERS_DIR / "SS27_заказы_по_категориям.xlsx"

# Canonical category labels as in procurement form (gender-specific).
MEN_CATS = [
    "Верхняя одежда муж",
    "Пиджаки, жакеты, бомбер муж",
    "Футболки, поло муж",
    "Брюки, джинсы муж",
    "Бриджи, шорты муж",
    "Трикотаж муж",
    "Рубашки",
    "Костюмы муж",
    "Обувь муж",
    "Аксессуары муж",
]
WOMEN_CATS = [
    "Верхняя одежда жен",
    "Пиджаки, жакеты, бомбер жен",
    "Футболки, поло, топы жен",
    "Блузки, рубашки жен",
    "Трикотаж жен",
    "Брюки, джинсы жен",
    "Бриджи, шорты жен",
    "Платья жен",
    "Юбки жен",
    "Обувь жен",
    "Аксессуары жен",
]

# Multilingual keyword rules → category key (without gender suffix).
# Order matters: first match wins.
CAT_KEYWORDS: list[tuple[str, list[str]]] = [
    (
        "обувь",
        [
            "sneaker",
            "sneakers",
            "scarpe",
            "scarpa",
            "shoe",
            "shoes",
            "loafer",
            "sandal",
            "espadrille",
            "boot",
            "footwear",
        ],
    ),
    (
        "аксессуары",
        [
            "hat",
            "bag",
            "belt",
            "scarf",
            "glove",
            "socks",
            "tie",
            "cravatta",
            "cintura",
            "borsa",
            "cappello",
            "portacasco",
            "accessory",
            "accessories",
            "guanti",
            "wallet",
            "umbrella",
            "baseball hat",
            "duffel",
            "backpack",
            "beauty case",
            " baseball",
            "trolley",
            "cabin bag",
        ],
    ),
    (
        "костюмы",
        ["suit", "anzug", "abito uomo", "costume formal", "two-piece"],
    ),
    (
        "платья",
        ["dress", "kleid", "abito", "abito donna", "jersey dress"],
    ),
    (
        "юбки",
        ["skirt", "rock", "gonna", "jupe"],
    ),
    (
        "бриджи/шорты",
        [
            "bermuda",
            "shorts",
            "short/bermuda",
            "hose short",
            "pants bermuda",
            "chino short",
            "short trousers",
            "short pants",
        ],
    ),
    (
        "верхняя",
        [
            "parka",
            "mantel",
            "trench",
            "raincoat",
            "outerwear",
            "padded multi",
            "windbreaker",
            "overcoat",
            "winter coat",
            "down jacket",
            "puffer",
        ],
    ),
    (
        "пиджаки",
        [
            "bomber",
            "blazer",
            "sakko",
            "jacket",
            "giacca",
            "giubbotto",
            "giubbino",
            "jacke",
            "waistcoat",
            "gilet",
            "overshirt",
            "outer jacket",
            "short jacket",
            "quilted",
            "cappuccio",
        ],
    ),
    (
        "футболки",
        [
            "t-shirt",
            "tshirt",
            "tee ",
            "polo",
            "top ",
            " top",
            "tank",
            "crewneck t-shirt",
            "maglietta",
            "g/c",  # Diktat girocollo cotton tops
            "girocollo",
        ],
    ),
    (
        "блузки/рубашки",
        [
            "blouse",
            "bluse",
            "camicia",
            "shirt",
            "hemd",
            "new kent",
            "business kent",
            "spread kent",
            "classic kent",
        ],
    ),
    (
        "брюки",
        [
            "trousers",
            "trouser",
            "pants",
            "pantaloni",
            "pantalone",
            "hose",
            "jeans",
            "chino",
            "5-pocket",
            "flatfront",
            "barrel fit",
            "wide fit",
        ],
    ),
    (
        "трикотаж",
        [
            "knit",
            "jumper",
            "sweater",
            "pullover",
            "cardigan",
            "felpa",
            "hoodie",
            "sweatshirt",
            "maglia",
            "strick",
            "jersey",
            "crewneck",
            "tuta",
        ],
    ),
]


def money(s: str | float | int | None) -> float:
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    t = str(s).strip()
    t = t.replace("€", "").replace("EUR", "").replace(" ", "").replace("\xa0", "")
    if not t:
        return 0.0
    # 9 255,10 or 9.255,10 or 9255.10 or 1,997.30
    if re.search(r",\d{2}$", t) and "." in t:
        # 1.997,30 or 5.848,00
        t = t.replace(".", "").replace(",", ".")
    elif re.search(r",\d{2}$", t):
        t = t.replace(",", ".")
    elif re.search(r"\.\d{2}$", t) and t.count(",") == 0:
        pass
    elif "," in t and "." in t:
        # 1,997.30 US
        t = t.replace(",", "")
    return float(t)


def categorize(text: str, gender: str) -> str:
    n = (text or "").lower()
    n = re.sub(r"\s+", " ", n)
    key = "прочее"
    for cat, kws in CAT_KEYWORDS:
        if any(k in n for k in kws):
            key = cat
            break
    return label_for(key, gender)


def label_for(key: str, gender: str) -> str:
    g = gender
    mapping_m = {
        "верхняя": "Верхняя одежда муж",
        "пиджаки": "Пиджаки, жакеты, бомбер муж",
        "футболки": "Футболки, поло муж",
        "брюки": "Брюки, джинсы муж",
        "бриджи/шорты": "Бриджи, шорты муж",
        "трикотаж": "Трикотаж муж",
        "блузки/рубашки": "Рубашки",
        "костюмы": "Костюмы муж",
        "обувь": "Обувь муж",
        "аксессуары": "Аксессуары муж",
        "платья": "Платья жен",  # shouldn't appear for men
        "юбки": "Юбки жен",
        "прочее": "Прочее",
    }
    mapping_w = {
        "верхняя": "Верхняя одежда жен",
        "пиджаки": "Пиджаки, жакеты, бомбер жен",
        "футболки": "Футболки, поло, топы жен",
        "брюки": "Брюки, джинсы жен",
        "бриджи/шорты": "Бриджи, шорты жен",
        "трикотаж": "Трикотаж жен",
        "блузки/рубашки": "Блузки, рубашки жен",
        "костюмы": "Костюмы муж",
        "обувь": "Обувь жен",
        "аксессуары": "Аксессуары жен",
        "платья": "Платья жен",
        "юбки": "Юбки жен",
        "прочее": "Прочее",
    }
    return (mapping_m if g == "муж" else mapping_w).get(key, "Прочее")


@dataclass
class Line:
    brand: str
    gender: str
    source: str
    name: str
    amount: float
    category: str = ""
    qty: float | None = None
    note: str = ""

    def finalize(self):
        if not self.category:
            self.category = categorize(self.name, self.gender)
        return self


@dataclass
class ParseResult:
    lines: list[Line] = field(default_factory=list)
    order_total_hint: float | None = None


def parse_filename(path: Path) -> tuple[str, str]:
    stem = path.stem
    # strip trailing " 2"
    stem = re.sub(r"\s+2$", "", stem)
    gender = "муж" if "муж" in stem else "жен"
    brand = re.sub(r"\s*(аксессуары\s*)?(муж|жен)\s*$", "", stem).strip()
    brand = re.sub(r"\s+", " ", brand)
    return brand, gender


def pdf_text(path: Path) -> str:
    r = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        capture_output=True,
        text=True,
    )
    return r.stdout or ""


def pdfplumber_text(path: Path) -> str:
    with pdfplumber.open(str(path)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


# ---------- brand parsers ----------


def parse_transit(path: Path, brand: str, gender: str) -> ParseResult:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    res = ParseResult()
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        model = cells[2] if len(cells) > 2 else None
        style = cells[1] if len(cells) > 1 else None
        qty = cells[13] if len(cells) > 13 else None
        price = cells[14] if len(cells) > 14 else None
        if model in (None, "Modello\nModel") or style is None:
            if cells[0] == "TOTAL" and price is not None:
                res.order_total_hint = money(price)
            continue
        if not isinstance(model, str):
            continue
        q = float(qty or 0)
        p = float(price or 0)
        amount = q * p
        name = f"{model} {style}"
        # waistcoat → трикотаж/жилет
        res.lines.append(
            Line(brand, gender, path.name, name, amount, qty=q).finalize()
        )
    return res


def parse_riani(path: Path, brand: str, gender: str) -> ParseResult:
    text = pdf_text(path)
    res = ParseResult()
    # article + type on following line(s), then pos line with total
    # 741750-2272\n               jacket
    articles = list(
        re.finditer(
            r"(\d{6}-\d{4})\s*\n\s*([a-z][a-z0-9 ()/'-]*(?:\n\s*[a-z][a-z0-9 ()/'-]*)?)",
            text,
            re.I,
        )
    )
    totals = list(
        re.finditer(
            r"^\s*\d+\s+\d+[^\n]*?(\d{1,3}(?:\.\d{3})*,\d{2})\s*$",
            text,
            re.M,
        )
    )
    # Pair by order: each article block followed by a total line
    # More robust: find type near each total by looking backward
    for m in totals:
        end = m.start()
        chunk = text[max(0, end - 400) : end]
        am = money(m.group(1))
        if am <= 0:
            continue
        type_m = re.search(
            r"(\d{6}-\d{4})\s*\n\s*([a-z][^\n]*(?:\n\s*[a-z][^\n]*)?)",
            chunk,
            re.I,
        )
        name = type_m.group(2).replace("\n", " ").strip() if type_m else chunk[-80:]
        # skip theme totals that are huge order totals duplicated
        if am >= 5000:
            res.order_total_hint = am
            continue
        res.lines.append(Line(brand, gender, path.name, name, am).finalize())
    # order total
    ot = re.search(r"Order Base Amount:\s*([\d.,]+)", text)
    if ot:
        res.order_total_hint = money(ot.group(1))
    return res


def parse_seidensticker(path: Path, brand: str, gender: str) -> ParseResult:
    text = pdf_text(path)
    res = ParseResult()
    # lines with Gesamt amount: ... Menge Preis Gesamt
    # 01.850034 New Kent ... 5 29,60 148,00
    for m in re.finditer(
        r"(\d{2}\.\d{6})\s+([^\n]+?)\s+(\d+)\s+(\d+,\d{2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})",
        text,
    ):
        art, name, qty, price, total = m.groups()
        # name may include color code digits — keep model words
        res.lines.append(
            Line(
                brand,
                gender,
                path.name,
                f"shirt {name}",
                money(total),
                category=label_for("блузки/рубашки", gender),
                qty=float(qty),
            )
        )
    # color continuation lines without article: "05 Beige ... 5 25,90 129,50"
    for m in re.finditer(
        r"^\s+\d{2}\s+[A-Za-zäöüÄÖÜß][^\n]*?\s+(\d+)\s+(\d+,\d{2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})",
        text,
        re.M,
    ):
        qty, price, total = m.groups()
        # avoid double-count: only if not already counted via article lines
        # these are extra colorways of previous article
        am = money(total)
        # Check if this exact total+qty already added from article pattern nearby — still add colorways
        # Article pattern already includes first colorway; continuation is additional.
        # Heuristic: if line starts with color code only, add as shirt.
        res.lines.append(
            Line(
                brand,
                gender,
                path.name,
                "shirt colorway",
                am,
                category=label_for("блузки/рубашки", gender),
                qty=float(qty),
            )
        )
    ot = re.search(r"Gesamt:\s+\d+\s+([\d.,]+)\s*EUR", text)
    if ot:
        res.order_total_hint = money(ot.group(1))
    return res


def parse_diktat(path: Path, brand: str, gender: str) -> ParseResult:
    text = pdf_text(path)
    res = ParseResult()
    for m in re.finditer(
        r"\d+\s+(DK\d+-\d+)\s+-\s+([^\n]+)\n[^\n]*?\n?\s+(\d+)\s+(\d+\.\d{2})\s+(\d+\.\d{2})",
        text,
    ):
        art, name, qty, price, total = m.groups()
        # G/C = girocollo → футболки/трикотаж; POLO → футболки; BOMBER/GIACCA → пиджаки
        res.lines.append(
            Line(brand, gender, path.name, f"{art} {name}", money(total), qty=float(qty)).finalize()
        )
    ot = re.search(r"Totale imponibile[^\d]*([\d,]+\.\d{2})", text)
    if ot:
        res.order_total_hint = money(ot.group(1))
    return res


def parse_gardeur(path: Path, brand: str, gender: str) -> ParseResult:
    """Gardeur: style-block subtotals = last line with one money value + small qty."""
    text = pdf_text(path)
    res = ParseResult()
    parts = re.split(r"(?m)(?=^\s*\d{1,2}\s{2,}[A-Z][A-Z0-9-]{2,})", text)
    for part in parts:
        m_head = re.match(r"\s*(\d{1,2})\s{2,}([A-Z][A-Z0-9-]*)", part)
        if not m_head:
            continue
        pos, style = int(m_head.group(1)), m_head.group(2)
        if pos > 40:
            continue
        type_bits = []
        for kw in [
            "Short/Bermuda",
            "Bermuda",
            "Hose",
            "Rock",
            "Jacket",
            "Blazer",
            "Shirt",
            "Top",
            "Denim",
        ]:
            if re.search(rf"\b{re.escape(kw)}\b", part, re.I):
                type_bits.append(kw)

        candidates = []
        for ln in part.splitlines():
            # skip grand totals with thousand-space: "3 072,00"
            if re.search(r"\d\s\d{3},\d{2}", ln):
                continue
            amounts = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", ln)
            if len(amounts) != 1:
                continue
            m = re.search(r"(\d{1,2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s*$", ln)
            if not m:
                continue
            qty = int(m.group(1))
            amount = money(m.group(2))
            if qty < 1 or qty > 40:
                continue
            # unit-price echo on size rows usually qty=1 and amount ~= unit
            if qty == 1 and amount < 90:
                continue
            candidates.append((qty, amount))
        if not candidates:
            continue
        qty, amount = candidates[-1]
        name = f"{style} " + " ".join(dict.fromkeys(type_bits))
        res.lines.append(
            Line(brand, gender, path.name, name.strip(), amount, qty=float(qty)).finalize()
        )

    spaced = re.findall(r"(\d(?:\s\d{3})+,\d{2})", text)
    if spaced:
        res.order_total_hint = max(money(x.replace(" ", "")) for x in spaced)
    else:
        plain = [money(x) for x in re.findall(r"(\d{3,},\d{2})", text)]
        if plain:
            res.order_total_hint = max(plain)
    return res

def parse_duno(path: Path, brand: str, gender: str) -> ParseResult:
    text = pdf_text(path)
    res = ParseResult()
    # Product blocks: STYLE - MODEL \n description \n CATEGORY_HEADER \n ... totals
    header_re = re.compile(
        r"(?m)^\s*((?:MAN|WOMAN)\s+[A-Z][A-Z0-9 /&-]{2,40}|HAT|BAG|BELT|TROLLEY)\s*$"
    )
    headers = list(header_re.finditer(text))
    money_re = re.compile(r"(\d+)\s+€([\d.]+,\d{2})\s+€([\d.]+,\d{2})")

    for i, hm in enumerate(headers):
        header = hm.group(1).strip()
        start = hm.end()
        end = headers[i + 1].start() if i + 1 < len(headers) else len(text)
        # description: look back up to 400 chars for a descriptive line
        back = text[max(0, hm.start() - 400) : hm.start()]
        desc = ""
        for ln in reversed(back.splitlines()):
            s = ln.strip()
            if not s or s.startswith("Quantity") or s.startswith("Proposal"):
                continue
            if re.match(r"^[A-Z0-9][A-Z0-9 /&-]{2,40}$", s) and s == s.upper():
                # style code line like "DOLPH - TREVI" — keep if no better
                if not desc:
                    desc = s
                continue
            if len(s) >= 8 and not s.startswith("€"):
                desc = s
                break
        chunk = text[start:end]
        for mm in money_re.finditer(chunk):
            qty, _price, total = mm.groups()
            am = money(total)
            if am <= 0:
                continue
            name = f"{desc} {header}".strip()
            res.lines.append(
                Line(brand, gender, path.name, name, am, qty=float(qty)).finalize()
            )

    ot = re.search(r"€([\d.]+,\d{2})\s+€0,00\s+\d+\s+€([\d.]+,\d{2})", text)
    if ot:
        res.order_total_hint = money(ot.group(2))
    return res


def parse_aeronautica(path: Path, brand: str, gender: str) -> ParseResult:
    text = pdfplumber_text(path)
    res = ParseResult()
    # Force category by file season line / filename
    forced = None
    if "аксессуар" in path.name.lower() or "ACCESSORIES" in text:
        forced = label_for("аксессуары", gender)
    elif "FOOTWEAR" in text or "SNEAKER" in text and "MAN" not in text.split("Season")[0][-200:]:
        if re.search(r"FOOTWEAR|SNEAKERS", text):
            # footwear order
            if "FOOTWEAR" in text:
                forced = label_for("обувь", gender)

    # Split by cod.
    parts = re.split(r"cod\.\s+", text)
    for part in parts[1:]:
        hm = re.match(r"(\S+)\s+([A-Z][A-Z0-9À-Ü /'-]{2,60})", part)
        if not hm:
            continue
        art, pname = hm.group(1), hm.group(2).strip()
        # color lines: COLOR ... qty €price €price
        for cm in re.finditer(
            r"([A-Z][A-Z0-9 /]+?)\s+(?:(?:\d\s+)+)?(\d+)\s+€([\d,]+)\s+€([\d,]+)",
            part,
        ):
            color, qty, price, _disc = cm.groups()
            if color.strip() in {"S", "M", "L", "XL"}:
                continue
            q = float(qty)
            unit = money(price)
            amount = q * unit
            name = f"{pname} {color.strip()}"
            cat = forced or categorize(pname, gender)
            res.lines.append(
                Line(brand, gender, path.name, name, amount, category=cat, qty=q)
            )
    return res


def _rr_group(sku: str) -> str:
    """Extract product-group code from Roy Robson SKU (2nd segment)."""
    s = sku.strip()
    if s.startswith(("D-", "S-")):
        s = "0-" + s[2:]
    if not s.startswith("0-"):
        # bare "02-18128-11949-00" → prepend 0-
        s = "0-" + s
    parts = s.split("-")
    return parts[1] if len(parts) > 1 else ""


def parse_roy_robson(path: Path, brand: str, gender: str) -> ParseResult:
    text = pdf_text(path)
    res = ParseResult()
    GROUP_MAP = {
        "00": "пиджаки",
        "07": "пиджаки",
        "02": "костюмы",
        "01": "брюки",
        "51": "брюки",
        "59": "блузки/рубашки",
        "76": "обувь",
        "90": "трикотаж",
        "91": "блузки/рубашки",
        "96": "трикотаж",
    }

    sku_re = re.compile(
        r"((?:[DS]-)?\d{2}-\d{5}-\d{5}-\d{2})"
    )
    # Split by SKU occurrences (each colorway / article)
    matches = list(sku_re.finditer(text))
    for i, m in enumerate(matches):
        sku = m.group(1)
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[m.start() : end]
        unit = None
        total_qty = 0
        for lm in re.finditer(
            r"EUR\s*([\d.,]+)\s+EUR\s*[\d.,]*\s+(\d+)\s",
            block,
        ):
            u, q = money(lm.group(1)), int(lm.group(2))
            if q > 0:
                unit = u
                total_qty += q
        if not unit or total_qty <= 0:
            continue
        group = _rr_group(sku)
        key = GROUP_MAP.get(group, "прочее")
        if group == "90":
            if unit <= 25:
                key = "футболки"
            elif unit <= 40:
                key = "блузки/рубашки"
            else:
                key = "трикотаж"
        res.lines.append(
            Line(
                brand,
                gender,
                path.name,
                f"RR {sku} g{group}",
                unit * total_qty,
                category=label_for(key, gender),
                qty=float(total_qty),
            )
        )

    ot = re.search(r"Total\s+\d+\s+([\d\s]+,\d{2})", text)
    if ot:
        res.order_total_hint = money(ot.group(1).replace(" ", ""))
    return res

def parse_file(path: Path) -> ParseResult:
    brand, gender = parse_filename(path)
    name = path.name.lower()
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return parse_transit(path, brand, gender)
    if "riani" in name:
        return parse_riani(path, brand, gender)
    if "seiden" in name:
        return parse_seidensticker(path, brand, gender)
    if "diktat" in name:
        return parse_diktat(path, brand, gender)
    if "gardeur" in name:
        return parse_gardeur(path, brand, gender)
    if "duno" in name:
        return parse_duno(path, brand, gender)
    if "aeronautica" in name:
        return parse_aeronautica(path, brand, gender)
    if "roy" in name or "robson" in name:
        return parse_roy_robson(path, brand, gender)
    # generic fallback
    return ParseResult()


def main():
    all_lines: list[Line] = []
    hints = {}
    files = sorted(ORDERS_DIR.iterdir(), key=lambda p: p.name.lower())
    for path in files:
        if path.suffix.lower() not in {".pdf", ".xlsx", ".xls"}:
            continue
        if path.name.startswith("SS27"):
            continue
        print(f"Parsing {path.name} ...")
        res = parse_file(path)
        print(f"  → {len(res.lines)} lines, sum={sum(l.amount for l in res.lines):.2f}, hint={res.order_total_hint}")
        all_lines.extend(res.lines)
        hints[path.name] = {
            "lines": len(res.lines),
            "sum": round(sum(l.amount for l in res.lines), 2),
            "hint": res.order_total_hint,
        }

    # Deduplicate Seidensticker if article+colorway double counted
    # (handled in parser carefully)

    # Aggregate brand+gender × category
    matrix: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    cats_used: set[str] = set()
    for ln in all_lines:
        row = f"{ln.brand} {ln.gender}"
        matrix[row][ln.category] += ln.amount
        cats_used.add(ln.category)

    # Column order: men cats then women cats then Прочее, only those used + always show common set
    col_order = []
    for c in MEN_CATS + WOMEN_CATS + ["Прочее"]:
        if c in cats_used and c not in col_order:
            col_order.append(c)
    for c in sorted(cats_used):
        if c not in col_order:
            col_order.append(c)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    header = ["Бренд"] + col_order + ["Итого"]
    ws.append(header)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    for col in range(1, len(header) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    rows_sorted = sorted(matrix.keys())
    for row_name in rows_sorted:
        vals = [matrix[row_name].get(c, 0.0) for c in col_order]
        total = sum(vals)
        ws.append([row_name] + [round(v, 2) if v else None for v in vals] + [round(total, 2)])

    # totals row
    grand = []
    for c in col_order:
        grand.append(round(sum(matrix[r].get(c, 0.0) for r in rows_sorted), 2))
    ws.append(["ИТОГО"] + [g or None for g in grand] + [round(sum(grand), 2)])
    for col in range(1, len(header) + 1):
        ws.cell(ws.max_row, col).font = Font(bold=True)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    for i in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Detail sheet
    ws2 = wb.create_sheet("Детали")
    ws2.append(
        ["Бренд", "Пол", "Файл", "Наименование", "Категория", "Кол-во", "Сумма EUR"]
    )
    for ln in all_lines:
        ws2.append(
            [
                ln.brand,
                ln.gender,
                ln.source,
                ln.name,
                ln.category,
                ln.qty,
                round(ln.amount, 2),
            ]
        )

    # Validation sheet
    ws3 = wb.create_sheet("Сверка")
    ws3.append(["Файл", "Сумма позиций", "Итог в документе", "Дельта"])
    for fname, h in sorted(hints.items()):
        hint = h["hint"]
        delta = None if hint is None else round(h["sum"] - hint, 2)
        ws3.append([fname, h["sum"], hint, delta])

    wb.save(OUT_XLSX)
    try:
        wb.save(OUT_XLSX_DL)
        print(f"Wrote {OUT_XLSX_DL}")
    except OSError as e:
        print(f"Could not write to Downloads ({e}), kept {OUT_XLSX}")
    OUT_DETAIL.write_text(
        json.dumps(
            {
                "hints": hints,
                "lines": [
                    {
                        "brand": l.brand,
                        "gender": l.gender,
                        "source": l.source,
                        "name": l.name,
                        "category": l.category,
                        "qty": l.qty,
                        "amount": round(l.amount, 2),
                    }
                    for l in all_lines
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"\nWrote {OUT_XLSX}")
    print(f"Wrote {OUT_DETAIL}")
    print(f"Grand total: {sum(l.amount for l in all_lines):.2f} EUR")


if __name__ == "__main__":
    main()
