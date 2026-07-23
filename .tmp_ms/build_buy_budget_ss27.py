# -*- coding: utf-8 -*-
"""Buy budget SS27 analytics for Antrasha — OTB + user method."""
import json
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AT = Path(
    "/Users/alekseiarsenev/.cursor/projects/"
    "Users-alekseiarsenev-WebstormProjects-antrasha-tinder/agent-tools"
)
OUT = Path(
    "/Users/alekseiarsenev/WebstormProjects/antrasha_tinder/"
    "Бюджет_закупки_ВЛ2027.xlsx"
)
SUMMARY = Path(
    "/Users/alekseiarsenev/WebstormProjects/antrasha_tinder/"
    ".tmp_ms/buy_budget_ss27_summary.json"
)


def load(name):
    return json.loads((AT / name).read_text())


profit_ss26 = load("647d9359-59cc-45b7-b471-cef61f34332e.txt")["items"]
profit_prior_full = load("278b0be5-ea58-4123-82f3-8209dc627543.txt")["items"]
profit_prior_early = load("0f512c56-e1e4-42b4-b318-39eda0c8543f.txt")["items"]
profit_prior_md = load("c7544613-5f0f-4b32-a888-ed02fb251c4c.txt")["items"]
def art_key(i):
    return i.get("article") or i.get("code") or i.get("name") or ""


w26 = {
    art_key(i): i
    for i in load("d51da5b9-7c58-4eab-abb8-c4bf0c0d192b.txt")["items"]
}
m26 = {
    art_key(i): i
    for i in load("ef28197e-7f13-490f-8a60-b9b83a5acce3.txt")["items"]
}
w25 = {
    art_key(i): i
    for i in load("ec25797d-bff7-46b5-9091-06664bc4adc0.txt")["items"]
}
m25 = {
    art_key(i): i
    for i in load("ec036f4b-fe7e-4acd-8872-fb6381224196.txt")["items"]
}

stock_prod = []
for f in [
    "04ec1373-b32a-43f7-a92d-01fc471a6e0d.txt",
    "e0ef33cf-7261-4c83-993f-0706b9d0efcb.txt",
    "33b79d2b-0d7a-4248-958c-f2199df5508a.txt",
]:
    stock_prod += load(f)["items"]

stock_var = []
for f in [
    "13c30fe6-112d-4926-bd48-6a89cc23a1f2.txt",
    "7b201b57-e8c4-45e8-8fcc-c4c8fb5b1353.txt",
    "001959d8-2e7b-42ff-abb4-9b23be9ce2b0.txt",
    "a18986c9-92b2-4f98-b8bf-600ee597f371.txt",
    "8191ee0a-9046-40e5-8ad3-70d53bfffc76.txt",
]:
    stock_var += load(f)["items"]

sales_26 = [
    ("2026-02", 1220290, 51),
    ("2026-03", 2201957, 88),
    ("2026-04", 3475907, 118),
    ("2026-05", 3892140.5, 124),
    ("2026-06", 3714989.2, 103),
    ("2026-07", 2392311, 80),
]
sales_25 = [
    ("2025-02", 1291280.5, 65),
    ("2025-03", 3417788.51, 105),
    ("2025-04", 3989449.8, 128),
    ("2025-05", 5217322.5, 149),
    ("2025-06", 4019191.95, 121),
    ("2025-07", 3833965.4, 132),
    ("2025-08", 2643031.3, 113),
]
sales_24 = [
    ("2024-02", 1575556, 69),
    ("2024-03", 3100672.5, 124),
    ("2024-04", 3668162.3, 108),
    ("2024-05", 3430563.2, 137),
    ("2024-06", 4449046, 140),
    ("2024-07", 3018527.5, 133),
    ("2024-08", 2922405.5, 142),
]

DATE_RE = re.compile(r"(?:^|/)(\d{2})\.(\d{2})(?:\b|$)")
SEASON_TAG_RE = re.compile(r"\((ВЛ|ОЗ)\s*20?(\d{2})", re.I)
SIZE_RE = re.compile(r"\((?:ВЛ|ОЗ)\s*20?\d{2},\s*([^)]+)\)", re.I)

FW_KW = [
    "куртк",
    "пальто",
    "пухов",
    "парка",
    "шуб",
    "дублен",
    "шапк",
    "шарф",
    "снуд",
    "манишк",
    "перчат",
    "вареж",
    "ботин",
    "сапог",
    "угг",
]
SS_KW = [
    "шорты",
    "сарафан",
    "купальн",
    "босонож",
    "сандал",
    "панам",
    "ветровк",
    "плащ",
    "тренч",
]
CAT_RULES = [
    (
        "Верхняя одежда ОЗ",
        ["куртк", "пальто", "пухов", "парка", "шуб", "дублен", "полупальт"],
    ),
    ("Верх SS (ветровки/плащи)", ["ветровк", "плащ", "тренч", "дождевик"]),
    ("Пиджаки/жакеты", ["пиджак", "жакет", "бомбер", "блейзер"]),
    ("Платья", ["платье", "сарафан"]),
    ("Юбки", ["юбка"]),
    ("Брюки/джинсы", ["брюк", "джинс", "бридж", "чинос"]),
    ("Шорты", ["шорт"]),
    ("Блузки/рубашки", ["блуз", "рубаш"]),
    ("Футболки/топы", ["футболк", "поло", "топ", "майк"]),
    (
        "Трикотаж",
        [
            "трикотаж",
            "пуловер",
            "джемпер",
            "свитер",
            "кардиган",
            "толстовк",
            "худи",
            "свитшот",
            "жилет",
            "водолаз",
        ],
    ),
    ("Костюмы", ["костюм", "комбинезон"]),
    (
        "Обувь",
        [
            "обув",
            "туфл",
            "ботин",
            "сапог",
            "кроссов",
            "лофер",
            "балетки",
            "босонож",
            "сандал",
            "мокасин",
            "слипон",
        ],
    ),
    (
        "Аксессуары",
        [
            "ремень",
            "пояс",
            "сумк",
            "кошел",
            "шарф",
            "шапк",
            "кепк",
            "бейсбол",
            "панам",
            "очк",
            "галстук",
            "бижут",
            "носк",
            "палантин",
            "зонт",
            "перчат",
            "вареж",
            "манишк",
            "снуд",
        ],
    ),
]


def gender_of(article, name, men_set, women_set):
    key = article or ""
    if key in women_set:
        return "Ж"
    if key in men_set:
        return "М"
    n = (name or "").lower()
    if "муж" in n:
        return "М"
    if "жен" in n:
        return "Ж"
    if any(k in n for k in ["блуз", "платье", "сарафан", "юбка", "топ"]):
        return "Ж"
    return "Н/Д"


def category_of(name):
    n = (name or "").lower()
    for cat, kws in CAT_RULES:
        if any(k in n for k in kws):
            return cat
    return "Прочее"


def season_of(name, article, code):
    n = name or ""
    m = SEASON_TAG_RE.search(n)
    if m:
        return "ВЛ" if m.group(1).upper() == "ВЛ" else "ОЗ"
    nl = n.lower()
    if any(k in nl for k in FW_KW):
        return "ОЗ"
    if any(k in nl for k in SS_KW):
        return "ВЛ"
    src = article or code or n
    dm = DATE_RE.search(src)
    if dm:
        mm = int(dm.group(1))
        if mm in (2, 3, 4, 5, 6, 7, 8):
            return "ВЛ"
        return "ОЗ"
    return "Н/Д"


def parse_date_tag(article, code, name):
    src = article or code or name or ""
    dm = DATE_RE.search(src)
    if not dm:
        return None
    return f"{dm.group(1)}.{dm.group(2)}"


def annotate(items, men_set, women_set):
    out = []
    for i in items:
        art = i.get("article") or i.get("code") or ""
        name = i.get("name") or ""
        out.append(
            {
                **i,
                "gender": gender_of(art, name, men_set, women_set),
                "season": season_of(name, art, i.get("code")),
                "category": category_of(name),
                "date_tag": parse_date_tag(art, i.get("code"), name),
            }
        )
    return out


ss26 = annotate(profit_ss26, m26, w26)
prior = annotate(profit_prior_full, m25, w25)
prior_early = annotate(profit_prior_early, m25, w25)
prior_md = annotate(profit_prior_md, m25, w25)


def agg(rows, keyfn):
    d = defaultdict(
        lambda: {
            "sellSum": 0,
            "sellCostSum": 0,
            "sellQuantity": 0,
            "profit": 0,
            "sku": 0,
        }
    )
    for r in rows:
        k = keyfn(r)
        d[k]["sellSum"] += r.get("sellSum") or 0
        d[k]["sellCostSum"] += r.get("sellCostSum") or 0
        d[k]["sellQuantity"] += r.get("sellQuantity") or 0
        d[k]["profit"] += r.get("profit") or 0
        d[k]["sku"] += 1
    return d


def annotate_stock_prod(items):
    out = []
    men_all = {**m26, **m25}
    women_all = {**w26, **w25}
    for i in items:
        art = i.get("article") or i.get("code") or ""
        name = i.get("name") or ""
        g = gender_of(art, name, men_all, women_all)
        s = season_of(name, art, i.get("code"))
        if "ВЛ" in name and s == "Н/Д":
            s = "ВЛ"
        if "ОЗ" in name and s == "Н/Д":
            s = "ОЗ"
        is_base = "антраша" in name.lower() and any(
            k in name.lower() for k in ["футболк", "поло", "топ"]
        )
        out.append(
            {
                **i,
                "gender": g,
                "season": s,
                "category": category_of(name),
                "is_base": is_base,
                "date_tag": parse_date_tag(art, i.get("code"), name),
            }
        )
    return out


stock_p = annotate_stock_prod(stock_prod)


def parent_article(v):
    art = v.get("article") or ""
    if art:
        return art
    name = v.get("name") or ""
    return re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()


def extract_size(name):
    m = SIZE_RE.search(name or "")
    if m:
        return m.group(1).strip()
    m2 = re.search(r",\s*([^)]+)\)\s*$", name or "")
    return m2.group(1).strip() if m2 else None


def season_tag_full(name):
    m = SEASON_TAG_RE.search(name or "")
    if not m:
        return None
    return ("ВЛ" if m.group(1).upper() == "ВЛ" else "ОЗ") + m.group(2)


styles = defaultdict(
    lambda: {
        "sizes": set(),
        "qty": 0,
        "cost": 0,
        "sale": 0,
        "gender": "Н/Д",
        "category": "Прочее",
        "season_tag": None,
    }
)
men_all = {**m26, **m25}
women_all = {**w26, **w25}
for v in stock_var:
    name = v.get("name") or ""
    tag = season_tag_full(name)
    art = parent_article(v)
    size = extract_size(name)
    st = styles[art]
    st["qty"] += v.get("stock") or 0
    st["cost"] += (v.get("price") or 0) * (v.get("stock") or 0)
    st["sale"] += (v.get("salePrice") or 0) * (v.get("stock") or 0)
    if size:
        st["sizes"].add(size)
    if tag:
        st["season_tag"] = tag
    st["category"] = category_of(name)
    st["gender"] = gender_of(art, name, men_all, women_all)


def totals(rows):
    return {
        "sellSum": sum(r["sellSum"] for r in rows),
        "sellCostSum": sum(r["sellCostSum"] for r in rows),
        "sellQuantity": sum(r["sellQuantity"] for r in rows),
        "profit": sum(r["profit"] for r in rows),
        "sku": len(rows),
    }


T26 = totals(ss26)
T25 = totals(prior)
T25e = totals(prior_early)
T25m = totals(prior_md)

by_season_26 = agg(ss26, lambda r: r["season"])
by_season_25 = agg(prior, lambda r: r["season"])

ss_share_26 = by_season_26["ВЛ"]["sellSum"] / T26["sellSum"] if T26["sellSum"] else 0
fw_share_26 = by_season_26["ОЗ"]["sellSum"] / T26["sellSum"] if T26["sellSum"] else 0

ss_only_26 = [r for r in ss26 if r["season"] == "ВЛ"]
ss_only_25 = [r for r in prior if r["season"] == "ВЛ"]
by_gc26 = agg(ss_only_26, lambda r: (r["gender"], r["category"]))
by_gc25 = agg(ss_only_25, lambda r: (r["gender"], r["category"]))
g_cost_26 = agg(ss_only_26, lambda r: r["gender"])
g25c = agg(ss_only_25, lambda r: r["gender"])

cogs_ratio_early = T25e["sellCostSum"] / T25e["sellSum"] if T25e["sellSum"] else 0
cogs_ratio_md = T25m["sellCostSum"] / T25m["sellSum"] if T25m["sellSum"] else 0
cogs_ratio_26 = T26["sellCostSum"] / T26["sellSum"] if T26["sellSum"] else 0

retail_25_full = sum(x[1] for x in sales_25)
retail_26_ytd = sum(x[1] for x in sales_26)
prior_jul_full = sales_25[5][1]
prior_aug = sales_25[6][1]
prior_remain = prior_jul_full * (8 / 31) + prior_aug
prior_ytd_to_jul23 = sum(x[1] for x in sales_25[:5]) + prior_jul_full * (23 / 31)
yoy_ytd = retail_26_ytd / prior_ytd_to_jul23 if prior_ytd_to_jul23 else 1
proj_remain_retail = prior_remain * yoy_ytd

ytd_retail_profit = T26["sellSum"]
scale = ytd_retail_profit / retail_26_ytd if retail_26_ytd else 1
proj_remain_retail_p = proj_remain_retail * scale
proj_full_retail_p = ytd_retail_profit + proj_remain_retail_p
proj_remain_cogs = proj_remain_retail_p * cogs_ratio_md
proj_full_cogs = T26["sellCostSum"] + proj_remain_cogs

ss_cogs_share = (
    by_season_26["ВЛ"]["sellCostSum"] / T26["sellCostSum"]
    if T26["sellCostSum"]
    else ss_share_26
)
proj_ss_cogs = proj_full_cogs * ss_cogs_share
proj_ss_retail = proj_full_retail_p * ss_share_26

avg_cost_26 = T26["sellCostSum"] / T26["sellQuantity"] if T26["sellQuantity"] else 0
avg_cost_25 = T25["sellCostSum"] / T25["sellQuantity"] if T25["sellQuantity"] else 0
avg_retail_26 = T26["sellSum"] / T26["sellQuantity"] if T26["sellQuantity"] else 0
avg_retail_25 = T25["sellSum"] / T25["sellQuantity"] if T25["sellQuantity"] else 0
price_inflation_cost = (avg_cost_26 / avg_cost_25 - 1) if avg_cost_25 else 0
price_inflation_retail = (avg_retail_26 / avg_retail_25 - 1) if avg_retail_25 else 0

qty_25_ytd = sum(x[2] for x in sales_25[:5]) + int(sales_25[5][2] * (23 / 31))
qty_25_remain = int(sales_25[5][2] * (8 / 31)) + sales_25[6][2]
qty_26_plot_ytd = sum(x[2] for x in sales_26)
yoy_qty = qty_26_plot_ytd / qty_25_ytd if qty_25_ytd else 1
proj_qty_full = qty_26_plot_ytd + qty_25_remain * yoy_qty

r24 = sum(x[1] for x in sales_24)
r26 = proj_full_retail_p
cagr = (r26 / r24) ** 0.5 - 1 if r24 else 0
growth = cagr * 0.5 if cagr > 0 else cagr

ss_stock = [s for s in stock_p if s["season"] == "ВЛ" and not s["is_base"]]
fw_stock = [s for s in stock_p if s["season"] == "ОЗ"]


def is_ss26_fresh(s):
    t = s.get("date_tag")
    if not t:
        return False
    mm, yy = t.split(".")
    return yy == "26" and int(mm) in (2, 3, 4, 5, 6, 7, 8)


fresh_ss = [s for s in ss_stock if is_ss26_fresh(s)]
old_ss = [s for s in ss_stock if not is_ss26_fresh(s)]


def stock_tot(rows):
    return {
        "sku": len(rows),
        "qty": sum(r.get("stock") or 0 for r in rows),
        "cost": sum((r.get("price") or 0) * (r.get("stock") or 0) for r in rows),
        "sale": sum((r.get("salePrice") or 0) * (r.get("stock") or 0) for r in rows),
    }


ST_ss = stock_tot(ss_stock)
ST_fw = stock_tot(fw_stock)
ST_fresh = stock_tot(fresh_ss)
ST_old = stock_tot(old_ss)
ST_all = stock_tot(stock_p)

CARRY_FRESH = 0.30
CARRY_OLD = 0.05
carry_cost = ST_fresh["cost"] * CARRY_FRESH + ST_old["cost"] * CARRY_OLD

weeks_cover_target = 8
weekly_ss_cogs = proj_ss_cogs / 28
target_bom_ss27 = weekly_ss_cogs * weeks_cover_target
planned_ss27_cogs_sales = (
    proj_ss_cogs * (1 + growth) * (1 + max(price_inflation_cost, 0))
)
target_eom_ss27 = weekly_ss_cogs * (1 + growth) * 6
otb_receipts = max(planned_ss27_cogs_sales + target_eom_ss27 - carry_cost, 0)
user_buy = max(planned_ss27_cogs_sales * 1.05 - carry_cost, 0)
buy_rub = (otb_receipts + user_buy) / 2

EUR_RATE = 95.0
LOGISTICS = 0.10
buy_eur = (buy_rub * (1 - LOGISTICS)) / EUR_RATE
avg_cost_buy = avg_cost_26 * (1 + max(price_inflation_cost, 0))
buy_units = buy_rub / avg_cost_buy if avg_cost_buy else 0

ss_cost_26 = sum(r["sellCostSum"] for r in ss_only_26) or 1
ss_cost_25 = sum(r["sellCostSum"] for r in ss_only_25) or 1
g_share = {g: g_cost_26[g]["sellCostSum"] / ss_cost_26 for g in g_cost_26}


def mix_weights(by_gc, total_cost):
    return {
        k: (v["sellCostSum"] / total_cost if total_cost else 0)
        for k, v in by_gc.items()
    }


w26m = mix_weights(by_gc26, ss_cost_26)
w25m = mix_weights(by_gc25, ss_cost_25)

vel = agg(ss_only_26, lambda r: (r["gender"], r["category"]))
st_gc = defaultdict(lambda: {"qty": 0, "cost": 0, "sku": 0})
for s in ss_stock:
    k = (s["gender"], s["category"])
    st_gc[k]["qty"] += s.get("stock") or 0
    st_gc[k]["cost"] += (s.get("price") or 0) * (s.get("stock") or 0)
    st_gc[k]["sku"] += 1

size_gc = defaultdict(
    lambda: {"styles": 0, "deep3": 0, "avg_sizes": 0, "sizes_sum": 0, "depth_pct": 0}
)
for art, st in styles.items():
    if not st["season_tag"] or not st["season_tag"].startswith("ВЛ"):
        continue
    k = (st["gender"], st["category"])
    n = len(st["sizes"])
    size_gc[k]["styles"] += 1
    size_gc[k]["sizes_sum"] += n
    if n >= 3:
        size_gc[k]["deep3"] += 1
for k, v in size_gc.items():
    v["avg_sizes"] = v["sizes_sum"] / v["styles"] if v["styles"] else 0
    v["depth_pct"] = v["deep3"] / v["styles"] if v["styles"] else 0


def adj_factor(g, cat):
    k = (g, cat)
    sales_qty = vel[k]["sellQuantity"] if k in vel else 0
    stock_qty = st_gc[k]["qty"]
    season_weeks = 22
    weekly = sales_qty / season_weeks if season_weeks else 0
    cover = (stock_qty / weekly) if weekly > 0 else (99 if stock_qty > 0 else 0)
    depth = size_gc[k]["depth_pct"] if k in size_gc else 0.5
    share_now = w26m.get(k, 0)
    share_prev = w25m.get(k, 0)
    dyn = share_now - share_prev
    f = 1.0
    f += max(min(dyn * 2, 0.25), -0.25)
    if cover > 20:
        f *= 0.75
    elif cover > 12:
        f *= 0.90
    elif cover < 4 and weekly > 0:
        f *= 1.15
    if depth < 0.35 and weekly > 0:
        f *= 1.10
    if depth > 0.7 and cover > 10:
        f *= 0.95
    return f, cover, depth, dyn, share_now


alloc_rows = []
for g in ["Ж", "М", "Н/Д"]:
    g_budget = buy_rub * g_share.get(g, 0)
    cats = sorted(
        {c for (gg, c) in by_gc26.keys() if gg == g},
        key=lambda c: -by_gc26[(g, c)]["sellCostSum"],
    )
    if not cats:
        continue
    raw = []
    for c in cats:
        base = g_budget * (
            by_gc26[(g, c)]["sellCostSum"] / g_cost_26[g]["sellCostSum"]
            if g_cost_26[g]["sellCostSum"]
            else 0
        )
        f, cover, depth, dyn, share = adj_factor(g, c)
        raw.append((c, base * f, base, f, cover, depth, dyn, share))
    sraw = sum(x[1] for x in raw) or 1
    for c, adj, base, f, cover, depth, dyn, share in raw:
        final = g_budget * (adj / sraw)
        units = final / avg_cost_buy if avg_cost_buy else 0
        eur = (final * (1 - LOGISTICS)) / EUR_RATE
        st = st_gc[(g, c)]
        sales = by_gc26.get(
            (g, c), {"sellSum": 0, "sellCostSum": 0, "sellQuantity": 0}
        )
        sales25 = by_gc25.get(
            (g, c), {"sellSum": 0, "sellCostSum": 0, "sellQuantity": 0}
        )
        alloc_rows.append(
            {
                "gender": g,
                "category": c,
                "buy_rub": final,
                "buy_eur": eur,
                "buy_units": units,
                "base_rub": base,
                "adj": f,
                "cover_w": cover,
                "depth_pct": depth,
                "share_delta": dyn,
                "share": share,
                "sales_sum": sales["sellSum"],
                "sales_cost": sales["sellCostSum"],
                "sales_qty": sales["sellQuantity"],
                "sales25_cost": sales25["sellCostSum"],
                "sales25_qty": sales25["sellQuantity"],
                "stock_qty": st["qty"],
                "stock_cost": st["cost"],
                "stock_sku": st["sku"],
            }
        )

# ---------- Excel ----------
wb = Workbook()
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
header_fill = PatternFill("solid", fgColor="1F2937")
header_font = Font(bold=True, color="FFFFFF", size=11)
money = "#,##0"
money2 = "#,##0.00"
pct = "0.0%"


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


ws = wb.active
ws.title = "0. Метод"
lines = [
    "Бюджет закупки Весна–Лето 2027 (Антраша)",
    "Срез данных МойСклад: 2026-07-23",
    "",
    "1. Ваш метод (как есть)",
    "• Сезон YTD с первой поставки ВЛ (~15.02) по сегодня; доля ВЛ vs ОЗ в продажах",
    "• Экстраполяция до конца августа с поправкой на уценку (COGS/выручка выше)",
    "• Если доля ВЛ ~70%+ — опираемся на неё для объёма закупки следующего ВЛ",
    "• Раскладка М/Ж → категории склада; контроль руб vs штуки (инфляция себестоимости)",
    "• Учёт остатков и глубины размерного ряда",
    "",
    "2. Классика fashion retail (Open-to-Buy)",
    "• Receipts (закуп) = Plan Sales@cost + Plan EOM − BOM − OnOrder",
    "• Plan Sales@cost = прогноз продаж сезона × (1+рост) × (1+инфляция себеса для единиц)",
    "• BOM ≈ ожидаемый переходящий ВЛ после распродажи",
    "• EOM target = N недель покрытия (старт сезона 8 нед, конец 6 нед)",
    "• Аллокация по contribution × динамике доли × cover × size-depth",
    "",
    "3. Что сделано в этом файле",
    "• Синтез: итог закупа = среднее(OTB receipts, ваш top-down)",
    "• EUR: (руб × 0.90) / 95 — курс 95 с транзакционными издержками, 10% резерв на логистику",
    "• Вкладки: Сводка, Сезонность, МЖ, Категории, Остатки/размеры, Помесячно, Допущения",
    "",
    "4. Ограничения",
    "• Сезон ВЛ/ОЗ: метка в названии, иначе месяц в артикуле (02–08=ВЛ), иначе тип товара",
    "• Пол: папки Женская/Мужская коллекция + эвристики по типу",
    "• Переходящий остаток: 30% свежего ВЛ26 + 5% старого ВЛ — калибруйте на вкладке Допущения",
]
for i, line in enumerate(lines, 1):
    ws.cell(i, 1, line)
    if i == 1:
        ws.cell(i, 1).font = Font(bold=True, size=14)
ws.column_dimensions["A"].width = 110

ws = wb.create_sheet("1. Сводка")
rows_sum = [
    ("Показатель", "Значение", "Комментарий"),
    ("Период YTD", "2026-02-15 — 2026-07-23", "Сезон ВЛ26 до среза"),
    ("Горизонт экстраполяции", "до 2026-08-31", "как в вашей логике"),
    ("Выручка YTD (прибыль по товарам)", round(T26["sellSum"]), "МойСклад"),
    ("Себестоимость продаж YTD", round(T26["sellCostSum"]), ""),
    ("Шт YTD", T26["sellQuantity"], ""),
    ("Маржа YTD %", T26["profit"] / T26["sellSum"] if T26["sellSum"] else 0, ""),
    ("COGS/выручка YTD", cogs_ratio_26, "сейчас уже ближе к уценке"),
    ("COGS/выручка LY early (до 15.07)", cogs_ratio_early, "полный прайс"),
    (
        "COGS/выручка LY markdown (15.07–31.08)",
        cogs_ratio_md,
        "для экстраполяции хвоста",
    ),
    ("Доля ВЛ в выручке YTD", ss_share_26, "цель ориентир ≥70%"),
    ("Доля ОЗ в выручке YTD", fw_share_26, ""),
    (
        "Прогноз выручки сезона до 31.08",
        round(proj_full_retail_p),
        "YoY run-rate × хвост LY",
    ),
    (
        "Прогноз COGS сезона до 31.08",
        round(proj_full_cogs),
        "хвост × markdown COGS ratio",
    ),
    ("Прогноз COGS именно ВЛ", round(proj_ss_cogs), "база спроса следующего ВЛ"),
    (
        "YoY выручка сезона (proj vs LY full)",
        proj_full_retail_p / retail_25_full - 1 if retail_25_full else 0,
        "",
    ),
    ("CAGR 2г сезона (24→26proj)", cagr, ""),
    ("Рост для плана SS27 (50% CAGR)", growth, "консервативно"),
    ("Инфляция себес./шт YoY", price_inflation_cost, "критично для единиц"),
    ("Инфляция розницы/шт YoY", price_inflation_retail, ""),
    ("Остаток ВЛ себест. всего", round(ST_ss["cost"]), "без базы Антраша-футболки"),
    ("в т.ч. свежий ВЛ26", round(ST_fresh["cost"]), ""),
    ("в т.ч. старый ВЛ", round(ST_old["cost"]), ""),
    (
        "Ожидаемый переходящий BOM SS27",
        round(carry_cost),
        "30% fresh + 5% old",
    ),
    ("OTB receipts (классика)", round(otb_receipts), "Sales+EOM−BOM"),
    ("Top-down (ваш метод+5% refresh)", round(user_buy), ""),
    ("РЕКОМЕНДУЕМЫЙ ЗАКУП ₽", round(buy_rub), "среднее OTB и top-down"),
    ("Закуп EUR (товары)", round(buy_eur, 1), "(₽×0.9)/95"),
    ("Ориентир шт (по ср.себесу)", round(buy_units), "с учётом инфляции себеса"),
    ("Курс EUR", EUR_RATE, "с транзакционными издержками"),
    ("Резерв логистики", LOGISTICS, "вычтен из руб перед делением"),
]
for r_i, row in enumerate(rows_sum, 1):
    for c_i, v in enumerate(row, 1):
        ws.cell(r_i, c_i, v)
style_header(ws, 1, 3)
for r in range(2, len(rows_sum) + 1):
    label = ws.cell(r, 1).value or ""
    if any(
        x in label
        for x in (
            "Доля",
            "YoY",
            "CAGR",
            "Рост",
            "Инфляция",
            "Резерв",
            "Маржа",
            "COGS/",
        )
    ):
        ws.cell(r, 2).number_format = pct
    elif isinstance(ws.cell(r, 2).value, (int, float)) and "шт" not in label.lower() and "Курс" not in label:
        ws.cell(r, 2).number_format = money
    if "РЕКОМЕНДУЕМЫЙ" in label:
        for c in range(1, 4):
            ws.cell(r, c).font = Font(bold=True, size=12)
            ws.cell(r, c).fill = PatternFill("solid", fgColor="DBEAFE")
autosize(ws, [42, 18, 48])

ws = wb.create_sheet("2. Сезонность ВЛ_ОЗ")
ws.append(
    [
        "Сезон",
        "Выручка YTD",
        "Себест. YTD",
        "Шт",
        "Доля выр. %",
        "Доля себес. %",
        "LY выручка full SS",
        "LY себест.",
        "LY шт",
    ]
)
style_header(ws, 1, 9)
for season in ["ВЛ", "ОЗ", "Н/Д"]:
    a = by_season_26[season]
    b = by_season_25[season]
    ws.append(
        [
            season,
            round(a["sellSum"]),
            round(a["sellCostSum"]),
            a["sellQuantity"],
            a["sellSum"] / T26["sellSum"] if T26["sellSum"] else 0,
            a["sellCostSum"] / T26["sellCostSum"] if T26["sellCostSum"] else 0,
            round(b["sellSum"]),
            round(b["sellCostSum"]),
            b["sellQuantity"],
        ]
    )
for r in range(2, 5):
    ws.cell(r, 5).number_format = pct
    ws.cell(r, 6).number_format = pct
    for c in (2, 3, 7, 8):
        ws.cell(r, c).number_format = money
ws.append([])
status = "OK" if ss_share_26 >= 0.70 else "НИЖЕ — проверьте классификацию/микс зала"
ws.append(
    [
        "Вывод",
        f"Доля ВЛ в выручке YTD = {ss_share_26:.1%}. Ориентир «норм» ≥70%: {status}",
    ]
)
autosize(ws, [10, 14, 14, 8, 12, 12, 16, 12, 10])

ws = wb.create_sheet("3. Муж_Жен")
ws.append(
    [
        "Пол",
        "Выручка ВЛ YTD",
        "Себест. ВЛ YTD",
        "Шт",
        "Доля себес. %",
        "LY себест. ВЛ",
        "Δ доли п.п.",
        "Бюджет закупа ₽",
        "Бюджет EUR",
        "Ориентир шт",
    ]
)
style_header(ws, 1, 10)
for g in ["Ж", "М", "Н/Д"]:
    a = g_cost_26[g]
    b = g25c[g]
    share = a["sellCostSum"] / ss_cost_26
    share25 = b["sellCostSum"] / ss_cost_25 if ss_cost_25 else 0
    br = buy_rub * share
    ws.append(
        [
            g,
            round(a["sellSum"]),
            round(a["sellCostSum"]),
            a["sellQuantity"],
            share,
            round(b["sellCostSum"]),
            share - share25,
            round(br),
            round((br * (1 - LOGISTICS)) / EUR_RATE, 1),
            round(br / avg_cost_buy if avg_cost_buy else 0),
        ]
    )
for r in range(2, 5):
    ws.cell(r, 5).number_format = pct
    ws.cell(r, 7).number_format = "0.0%"
    for c in (2, 3, 6, 8):
        ws.cell(r, c).number_format = money
autosize(ws, [8, 16, 16, 8, 12, 14, 12, 14, 12, 12])

ws = wb.create_sheet("4. Категории закуп")
headers = [
    "Пол",
    "Категория",
    "Закуп ₽",
    "Закуп EUR",
    "Шт ориентир",
    "Доля ВЛ YTD",
    "Δ доли vs LY",
    "Adj фактор",
    "Cover нед.",
    "Depth ≥3разм %",
    "Продажи ₽ YTD",
    "Продажи себес.",
    "Продажи шт",
    "LY себес.",
    "Остаток шт",
    "Остаток себес.",
    "SKU остаток",
    "Сигнал",
]
ws.append(headers)
style_header(ws, 1, len(headers))
alloc_rows.sort(
    key=lambda x: (
        -(0 if x["gender"] == "Ж" else 1 if x["gender"] == "М" else 2),
        -x["buy_rub"],
    )
)
for a in alloc_rows:
    signal = []
    if a["adj"] >= 1.1:
        signal.append("нарастить")
    if a["adj"] <= 0.85:
        signal.append("урезать")
    if a["cover_w"] > 20:
        signal.append("много остатка")
    if a["cover_w"] < 4 and a["sales_qty"] > 0:
        signal.append("дыра в наличии")
    if a["depth_pct"] < 0.35 and a["sales_qty"] > 0:
        signal.append("ломаный размерный ряд")
    if a["share_delta"] > 0.02:
        signal.append("растёт")
    if a["share_delta"] < -0.02:
        signal.append("падает")
    ws.append(
        [
            a["gender"],
            a["category"],
            round(a["buy_rub"]),
            round(a["buy_eur"], 1),
            round(a["buy_units"]),
            a["share"],
            a["share_delta"],
            round(a["adj"], 2),
            round(a["cover_w"], 1),
            a["depth_pct"],
            round(a["sales_sum"]),
            round(a["sales_cost"]),
            a["sales_qty"],
            round(a["sales25_cost"]),
            a["stock_qty"],
            round(a["stock_cost"]),
            a["stock_sku"],
            ", ".join(signal) or "держать",
        ]
    )
for r in range(2, len(alloc_rows) + 2):
    for c in (3, 11, 12, 14, 16):
        ws.cell(r, c).number_format = money
    ws.cell(r, 4).number_format = money2
    for c in (6, 7, 10):
        ws.cell(r, c).number_format = pct
ws.append(
    [
        "ИТОГО",
        "",
        round(sum(a["buy_rub"] for a in alloc_rows)),
        round(sum(a["buy_eur"] for a in alloc_rows), 1),
        round(sum(a["buy_units"] for a in alloc_rows)),
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
)
last = len(alloc_rows) + 2
for c in range(1, 5):
    ws.cell(last, c).font = Font(bold=True)
    ws.cell(last, c).fill = PatternFill("solid", fgColor="DBEAFE")
autosize(
    ws,
    [6, 26, 12, 11, 11, 11, 11, 10, 10, 12, 12, 12, 10, 11, 10, 12, 10, 28],
)

ws = wb.create_sheet("5. Остатки_размеры")
ws.append(
    [
        "Пол",
        "Категория",
        "SKU ВЛ",
        "Шт",
        "Себест.",
        "Розница оценка",
        "Ср. размеров/стиль",
        "% стилей ≥3 размера",
        "Cover нед.",
        "Комментарий",
    ]
)
style_header(ws, 1, 10)
keys = sorted(set(list(st_gc.keys()) + list(size_gc.keys())))
for g, c in keys:
    st = st_gc[(g, c)]
    sg = size_gc[(g, c)]
    sales_qty = vel[(g, c)]["sellQuantity"] if (g, c) in vel else 0
    weekly = sales_qty / 22 if sales_qty else 0
    cover = st["qty"] / weekly if weekly else None
    avg_sz = sg["avg_sizes"] if sg["styles"] else None
    depth = sg["depth_pct"] if sg["styles"] else None
    comments = []
    if cover and cover > 20:
        comments.append("перезапас")
    elif cover and cover < 4:
        comments.append("нужна глубина/поставка")
    if depth is not None and depth < 0.35:
        comments.append("ломаный ряд")
    retail_est = sum(
        (x.get("salePrice") or 0) * (x.get("stock") or 0)
        for x in ss_stock
        if x["gender"] == g and x["category"] == c
    )
    ws.append(
        [
            g,
            c,
            st["sku"],
            st["qty"],
            round(st["cost"]),
            round(retail_est),
            round(avg_sz, 2) if avg_sz is not None else None,
            depth if depth is not None else None,
            round(cover, 1) if cover is not None else None,
            ", ".join(comments),
        ]
    )
for r in range(2, ws.max_row + 1):
    for c in (5, 6):
        ws.cell(r, c).number_format = money
    if ws.cell(r, 8).value is not None:
        ws.cell(r, 8).number_format = pct
autosize(ws, [6, 26, 10, 8, 12, 14, 14, 16, 10, 24])

ws2 = wb.create_sheet("5b. Сводка остатков")
ws2.append(["Сегмент", "SKU", "Шт", "Себест.", "Розница"])
style_header(ws2, 1, 5)
for label, st in [
    ("Все положительные", ST_all),
    ("ВЛ (без базы Антраша)", ST_ss),
    ("ВЛ26 свежий", ST_fresh),
    ("ВЛ старый", ST_old),
    ("ОЗ", ST_fw),
]:
    ws2.append([label, st["sku"], st["qty"], round(st["cost"]), round(st["sale"])])
for r in range(2, 7):
    for c in (4, 5):
        ws2.cell(r, c).number_format = money
autosize(ws2, [28, 8, 8, 14, 14])

ws = wb.create_sheet("6. Помесячно")
ws.append(
    [
        "Месяц",
        "Выручка 2026",
        "Шт 2026",
        "Выручка 2025",
        "Шт 2025",
        "Выручка 2024",
        "Шт 2024",
        "YoY 26/25",
    ]
)
style_header(ws, 1, 8)
months = ["02", "03", "04", "05", "06", "07", "08"]
map26 = {m[0][-2:]: m for m in sales_26}
map25 = {m[0][-2:]: m for m in sales_25}
map24 = {m[0][-2:]: m for m in sales_24}
for m in months:
    a = map26.get(m)
    b = map25.get(m)
    c = map24.get(m)
    yoy = (a[1] / b[1] - 1) if a and b and b[1] else None
    ws.append(
        [
            m,
            a[1] if a else None,
            a[2] if a else None,
            b[1] if b else None,
            b[2] if b else None,
            c[1] if c else None,
            c[2] if c else None,
            yoy,
        ]
    )
ws.append(
    [
        "Прогноз остатка Jul24–Aug31",
        round(proj_remain_retail_p),
        None,
        round(prior_remain),
        None,
        None,
        None,
        None,
    ]
)
ws.append(
    [
        "Итого сезон (proj)",
        round(proj_full_retail_p),
        round(proj_qty_full),
        round(retail_25_full),
        None,
        round(r24),
        None,
        proj_full_retail_p / retail_25_full - 1,
    ]
)
for r in range(2, ws.max_row + 1):
    for c in (2, 4, 6):
        if ws.cell(r, c).value is not None:
            ws.cell(r, c).number_format = money
    if ws.cell(r, 8).value is not None:
        ws.cell(r, 8).number_format = pct
autosize(ws, [28, 14, 10, 14, 10, 14, 10, 10])

ws = wb.create_sheet("7. Допущения")
assumptions = [
    ("Параметр", "Значение", "Зачем"),
    ("Сезон окно", "15.02–31.08", ""),
    ("Carry fresh SS26", CARRY_FRESH, "доля свежего ВЛ, доживёт до ВЛ27"),
    ("Carry old SS", CARRY_OLD, "доля старого ВЛ"),
    ("Target BOM cover weeks", weeks_cover_target, "покрытие на старт сезона"),
    ("Target EOM cover weeks", 6, "на конец сезона"),
    ("Growth vs proj SS26", growth, "50% от 2y CAGR"),
    (
        "Price inflation applied",
        max(price_inflation_cost, 0),
        "чтобы не потерять штуки",
    ),
    ("EUR rate", EUR_RATE, "с транзакционными"),
    ("Logistics reserve", LOGISTICS, "минус из рубль-бюджета"),
    ("Refresh premium top-down", 0.05, "в user_buy"),
]
for r in assumptions:
    ws.append(list(r))
style_header(ws, 1, 3)
for r in range(2, 12):
    if isinstance(ws.cell(r, 2).value, float) and ws.cell(r, 2).value <= 1:
        ws.cell(r, 2).number_format = pct
autosize(ws, [28, 14, 60])

ws = wb.create_sheet("8. Как читать")
txt = [
    "Сигналы на вкладке «4. Категории закуп»:",
    "• нарастить — растущая доля / низкий cover / ломаный размерный ряд",
    "• урезать — падающая доля или перезапас (cover >20 недель)",
    "• много остатка — сначала дожать уценкой, не дублировать закуп",
    "• дыра в наличии — приоритет поставки / глубины",
    "• ломаный размерный ряд — закупать с полным size curve, не «по 1 шт»",
    "",
    "EUR: сумма к заказу фабрикам/агентам по товару. Логистика 10% уже вычтена — везите отдельно из резерва.",
    "Сверяйте «Закуп ₽» с фактом контрактов; при росте курса >95 пересчитайте.",
]
for i, t in enumerate(txt, 1):
    ws.cell(i, 1, t)
ws.column_dimensions["A"].width = 100

wb.save(OUT)

summary = {
    "as_of": "2026-07-23",
    "ytd_retail": T26["sellSum"],
    "ytd_cogs": T26["sellCostSum"],
    "ytd_qty": T26["sellQuantity"],
    "ss_share": ss_share_26,
    "fw_share": fw_share_26,
    "proj_full_retail": proj_full_retail_p,
    "proj_full_cogs": proj_full_cogs,
    "proj_ss_cogs": proj_ss_cogs,
    "proj_ss_retail": proj_ss_retail,
    "cogs_ratio_26": cogs_ratio_26,
    "cogs_ratio_early": cogs_ratio_early,
    "cogs_ratio_md": cogs_ratio_md,
    "yoy_season": proj_full_retail_p / retail_25_full - 1,
    "cagr": cagr,
    "growth": growth,
    "infl_cost": price_inflation_cost,
    "infl_retail": price_inflation_retail,
    "avg_cost_26": avg_cost_26,
    "avg_cost_25": avg_cost_25,
    "stock_ss_cost": ST_ss["cost"],
    "stock_fresh": ST_fresh["cost"],
    "stock_old": ST_old["cost"],
    "carry": carry_cost,
    "otb": otb_receipts,
    "user_buy": user_buy,
    "buy_rub": buy_rub,
    "buy_eur": buy_eur,
    "buy_units": buy_units,
    "gender": {
        g: {
            "share": g_share.get(g, 0),
            "buy": buy_rub * g_share.get(g, 0),
            "eur": (buy_rub * g_share.get(g, 0) * (1 - LOGISTICS)) / EUR_RATE,
        }
        for g in ["Ж", "М", "Н/Д"]
    },
    "monthly_26": sales_26,
    "monthly_25": sales_25,
    "alloc_top": sorted(alloc_rows, key=lambda x: -x["buy_rub"])[:12],
    "season_ok": ss_share_26 >= 0.70,
}
SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=float))
print("Saved", OUT)
print("Buy RUB", round(buy_rub), "EUR", round(buy_eur), "units", round(buy_units))
print(
    "SS share",
    round(ss_share_26, 3),
    "YoY season",
    round(summary["yoy_season"], 3),
    "infl cost",
    round(price_inflation_cost, 3),
)
print("Gender shares", {k: round(v, 3) for k, v in g_share.items()})
print("Top alloc:")
for a in summary["alloc_top"][:8]:
    print(
        " ",
        a["gender"],
        a["category"],
        round(a["buy_rub"]),
        "EUR",
        round(a["buy_eur"]),
        "adj",
        round(a["adj"], 2),
        "cover",
        round(a["cover_w"], 1),
    )
